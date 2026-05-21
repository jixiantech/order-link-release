"""
gui.py - 苹果订单管理系统 GUI
"""

import sys
import json
import random
import string
import threading
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QLabel, QLineEdit, QTabWidget, QFileDialog, QMessageBox,
    QProgressBar, QSpinBox, QGroupBox, QFormLayout, QComboBox,
    QStatusBar, QSplitter, QTextEdit, QCheckBox, QDialog,
    QDialogButtonBox, QFrame
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QColor, QFont, QIcon

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── 版本信息 ──────────────────────────────────────────────────────
APP_VERSION = "1.0.3"
VERSION_URL = "https://raw.githubusercontent.com/jixiantech/order-link-release/main/version.txt"
DOWNLOAD_URL = "https://github.com/jixiantech/order-link-release/releases/latest/download/极限link.zip"

# ── 列定义 ────────────────────────────────────────────────────────
COLUMNS = [
    ("xc_link",     "XC 链接"),
    ("order_number","订单号"),
    ("order_email", "订单邮箱"),
    ("apple_id",    "苹果 ID"),
    ("apple_pwd",   "苹果密码"),
    ("phone",       "手机号"),
    ("new_address", "修改地址"),
    ("proxy_ip",    "代理 IP"),
    ("status",      "状态"),
    ("note",        "备注"),
]
COL_IDX = {k: i for i, (k, _) in enumerate(COLUMNS)}

STATUS_COLORS = {
    "待处理":   "#E8E8E8",
    "处理中":   "#FFF3CD",
    "绑定成功": "#D4EDDA",
    "取消成功": "#D4EDDA",
    "地址修改成功": "#D4EDDA",
    "失败":     "#F8D7DA",
    "跳过":     "#E2E3E5",
}

# ── Query Worker ─────────────────────────────────────────────────
class QueryWorker(QThread):
    row_ready = pyqtSignal(dict)
    finished = pyqtSignal()

    def __init__(self, xc_links, proxy_cfg):
        super().__init__()
        self.xc_links = xc_links
        self.proxy_cfg = proxy_cfg

    def _build_proxy_url(self, sid=None):
        cfg = self.proxy_cfg
        if not cfg.get("enabled") or not cfg.get("host"):
            return None
        import random, string
        if not sid:
            sid = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
        host = cfg["host"]
        port = cfg.get("port", "3120")
        user = cfg.get("user", "")
        pwd  = cfg.get("pwd", "")
        proxy_type = cfg.get("type", "Smartproxy")
        if user and pwd:
            if proxy_type == "Bright Data":
                return f"http://{user}-session-{sid}:{pwd}@{host}:{port}"
            else:
                return f"http://{user}_session-{sid}:{pwd}@{host}:{port}"
        return f"http://{host}:{port}"

    def make_proxy(self):
        """Race 机制：同时测试3个代理，选最快的"""
        import time, threading, requests as _req, warnings
        warnings.filterwarnings('ignore')
        cfg = self.proxy_cfg
        if not cfg.get("enabled") or not cfg.get("host"):
            return None
        candidates = [self._build_proxy_url() for _ in range(3)]
        results = []
        lock = threading.Lock()
        def test_proxy(proxy_url):
            try:
                start = time.time()
                r = _req.get("https://www.apple.com/xc/us/vieworder/",
                             proxies={"http": proxy_url, "https": proxy_url},
                             timeout=5, verify=False, allow_redirects=False)
                elapsed = time.time() - start
                if r.status_code < 500:
                    with lock:
                        results.append((elapsed, proxy_url))
            except Exception:
                pass
        threads = [threading.Thread(target=test_proxy, args=(p,)) for p in candidates]
        for t in threads: t.start()
        for t in threads: t.join(timeout=6)
        if results:
            results.sort(key=lambda x: x[0])
            return results[0][1]
        return candidates[0]

    def run(self):
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(self.xc_links)) as ex:
            futures = [ex.submit(self.query_one, link) for link in self.xc_links]
            for f in concurrent.futures.as_completed(futures):
                try:
                    result = f.result()
                    self.row_ready.emit(result)
                except Exception as e:
                    self.row_ready.emit({"xc_link": "", "order_number": "", 
                                         "order_status": "查询失败", "order_date": "",
                                         "device": str(e), "price": ""})
        self.finished.emit()

    def query_one(self, xc_link):
        import re, requests, json, warnings
        from bs4 import BeautifulSoup
        warnings.filterwarnings('ignore')

        result = {"xc_link": xc_link, "order_number": "", "order_email": "",
                  "order_status": "",
                  "order_date": "", "device": "", "tracking": "",
                  "carrier": "", "eta": "", "address": ""}

        # 从 xc 链接解析订单号和邮箱
        m = re.search(r'/vieworder/([^/]+)/([^/\s?]+)', xc_link)
        if not m:
            result["order_status"] = "链接无效"
            return result
        order_number = m.group(1)
        order_email = m.group(2)
        result["order_number"] = order_number
        result["order_email"] = order_email

        # 根据代理设置决定是否用代理
        proxy = self.make_proxy()
        try:
            from curl_cffi import requests as curl_requests
            import random
            _impersonate = random.choice([
                "chrome107", "chrome110", "chrome116", "chrome119",
                "chrome120", "chrome123", "chrome124", "chrome131",
                "chrome136", "chrome142"
            ])
            session = curl_requests.Session(impersonate=_impersonate, http_version=1)
        except ImportError:
            import requests as _req
            session = _req.Session()
        session.verify = False
        if proxy:
            session.proxies = {"http": proxy, "https": proxy}
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9",
        })
        try:
            import sys
            ip = session.get("https://api.ipify.org", timeout=5, verify=False).text.strip()
            print(f"[Query] IP: {ip}", file=sys.stderr)
        except Exception as e:
            print(f"[Query] IP获取失败: {e}", file=sys.stderr)

        try:
            # Step 1: xc → guest URL
            r = session.get(xc_link, allow_redirects=False, timeout=15)
            location = r.headers.get("location", "")
            if not location:
                result["order_status"] = "链接无效或已过期"
                return result

            # 解析 secure_base 和 token
            sb_m = re.match(r'(https://secure[\w]*\.store\.apple\.com)/shop/order/guest/[^/]+/([a-f0-9]+)', location)
            secure_base = sb_m.group(1) if sb_m else "https://secure.store.apple.com"
            token = sb_m.group(2) if sb_m else ""

            # 跟进重定向到 guest 页面
            if secure_base == "https://secure.store.apple.com":
                r2 = session.get(location, allow_redirects=False, timeout=15)
                loc2 = r2.headers.get("location", location)
                sb_m2 = re.match(r'(https://secure[\w]+\.store\.apple\.com)', loc2)
                if sb_m2:
                    secure_base = sb_m2.group(1)
                location = loc2 if loc2 else location

            # Step 2: GET guest 页面
            r = session.get(location + ("" if "?" in location else "?e=true"),
                          allow_redirects=True, timeout=15)

            # 解析 init_data
            soup = BeautifulSoup(r.text, 'html.parser')
            el = soup.find('script', id='init_data')
            if not el:
                result["order_status"] = "无法获取订单信息"
                return result

            init_data = json.loads(el.string)
            order_detail = init_data.get("orderDetail", {})

            # 订单状态
            order_items_keys = order_detail.get("orderItems", {}).get("c", [])
            if order_items_keys:
                first_item = order_detail.get("orderItems", {}).get(order_items_keys[0], {})
                status_tracker = first_item.get("orderItemStatusTracker", {}).get("d", {})
                current_status = status_tracker.get("currentStatus", "")
                result["order_status"] = current_status or "未知"

                # 购买设备
                item_details = first_item.get("orderItemDetails", {}).get("d", {})
                device = item_details.get("productName", "") or item_details.get("productDescription", "")
                result["device"] = device

                # 金额
                price = item_details.get("unitPrice", "") or item_details.get("totalPrice", "")
                result["price"] = price

            # 下单日期
            order_header = order_detail.get("orderHeader", {}).get("d", {})
            result["order_date"] = order_header.get("orderPlacedDate", "")

            # 物流号、物流公司、预计到达、地址
            tracking = ""
            carrier = ""
            eta = ""
            address = ""
            if order_items_keys:
                first_item = order_detail.get("orderItems", {}).get(order_items_keys[0], {})
                item_d = first_item.get("orderItemDetails", {}).get("d", {})

                # 物流号
                tracking_url_map = item_d.get("trackingURLMap", {})
                if tracking_url_map:
                    tracking = ", ".join(tracking_url_map.keys())

                # 预计到达
                eta = item_d.get("deliveryDate", "") or item_d.get("estimatedDeliveryDate", "")

                # 物流公司 - 从 trackingURL 判断
                for url in tracking_url_map.values():
                    if "fedex" in url.lower():
                        carrier = "FedEx"
                    elif "ups" in url.lower():
                        carrier = "UPS"
                    elif "usps" in url.lower():
                        carrier = "USPS"
                    elif "dhl" in url.lower():
                        carrier = "DHL"
                    break

                # 收货地址 from shippingInfo.shipping-address.address.d
                import sys
                shipping_addr = (first_item.get("shippingInfo", {})
                                .get("shipping-address", {})
                                .get("address", {})
                                .get("d", {}))
                print(f"[Query] shipping-address.d: {shipping_addr}", file=sys.stderr)
                if shipping_addr:
                    first_name = shipping_addr.get("firstName", "").strip()
                    last_name = shipping_addr.get("lastName", "").strip()
                    street = shipping_addr.get("street", "") or shipping_addr.get("address1", "")
                    city = shipping_addr.get("city", "")
                    state = shipping_addr.get("state", "")
                    postal = shipping_addr.get("postalCode", "") or shipping_addr.get("zip", "")
                    parts = [p for p in [f"{first_name} {last_name}".strip(), street, city, state, postal] if p]
                    address = ", ".join(parts)

            result["tracking"] = tracking
            result["carrier"] = carrier
            result["eta"] = eta
            result["address"] = address

            # 如果 orderDetail 为空，试试其他路径
            if not result["order_status"]:
                result["order_status"] = "已获取"

        except Exception as e:
            import sys
            print(f"[Query] 查询失败详情: {e}", file=sys.stderr)
            result["order_status"] = f"查询失败: {str(e)[:30]}"

        return result


# ── Worker Thread ─────────────────────────────────────────────────
class WorkerThread(QThread):
    row_update = pyqtSignal(int, str, str)  # row, status, note
    proxy_ip_signal = pyqtSignal(int, str)   # row, ip
    log_signal = pyqtSignal(str)
    finished = pyqtSignal()

    def __init__(self, task, rows, proxy_cfg, concurrency, new_address=None):
        super().__init__()
        self.task = task        # "link" / "cancel" / "address"
        self.rows = rows        # list of (row_idx, row_data)
        self.proxy_cfg = proxy_cfg
        self.concurrency = concurrency
        self.new_address = new_address or {}
        self._stop = False

    def stop(self):
        self._stop = True
        self._executor_shutdown = True

    def _build_proxy_url(self, sid=None):
        """根据代理类型生成代理 URL"""
        cfg = self.proxy_cfg
        if not cfg.get("enabled") or not cfg.get("host"):
            return None
        if not sid:
            sid = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
        host = cfg["host"]
        port = cfg.get("port", "3120")
        user = cfg.get("user", "")
        pwd  = cfg.get("pwd", "")
        proxy_type = cfg.get("type", "Smartproxy")
        if user and pwd:
            if proxy_type == "Bright Data":
                return f"http://{user}-session-{sid}:{pwd}@{host}:{port}"
            else:
                return f"http://{user}_session-{sid}:{pwd}@{host}:{port}"
        return f"http://{host}:{port}"

    def make_proxy(self):
        """Race 机制：同时测试3个代理，选最快的"""
        import time
        import requests as _req
        import warnings
        warnings.filterwarnings('ignore')

        cfg = self.proxy_cfg
        if not cfg.get("enabled") or not cfg.get("host"):
            return None

        # 生成3个候选代理
        candidates = [self._build_proxy_url() for _ in range(3)]

        results = []
        import threading
        lock = threading.Lock()

        def test_proxy(proxy_url):
            try:
                start = time.time()
                r = _req.get("https://www.apple.com/xc/us/vieworder/",
                             proxies={"http": proxy_url, "https": proxy_url},
                             timeout=5, verify=False, allow_redirects=False)
                elapsed = time.time() - start
                if r.status_code < 500:
                    with lock:
                        results.append((elapsed, proxy_url))
            except Exception:
                pass

        threads = [threading.Thread(target=test_proxy, args=(p,)) for p in candidates]
        for t in threads: t.start()
        for t in threads: t.join(timeout=6)

        if results:
            results.sort(key=lambda x: x[0])
            fastest = results[0]
            self.log_signal.emit(f"Race 代理选择: 延迟 {int(fastest[0]*1000)}ms")
            return fastest[1]

        # 全部失败则随机返回一个
        return candidates[0]

    def run(self):
        import concurrent.futures
        self._executor_shutdown = False
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.concurrency) as ex:
            futures = {}
            for row_idx, row_data in self.rows:
                if self._stop:
                    self.row_update.emit(row_idx, "已停止", "")
                    break
                self.row_update.emit(row_idx, "处理中", "")
                f = ex.submit(self.process_one, row_idx, row_data)
                futures[f] = row_idx
            for f in concurrent.futures.as_completed(futures):
                row_idx = futures[f]
                if self._stop:
                    self.row_update.emit(row_idx, "已停止", "")
                    f.cancel()
                    continue
                try:
                    status, note = f.result(timeout=1)
                except concurrent.futures.TimeoutError:
                    status, note = "已停止", ""
                except Exception as e:
                    status, note = "失败", str(e)
                self.row_update.emit(row_idx, status, note)
        self.finished.emit()

    def _friendly_error(self, err):
        """把技术性错误转成普通人能看懂的中文"""
        err = str(err)
        if "HTTPSConnectionPool" in err or "ProxyError" in err or "ConnectionError" in err:
            return "网络连接失败，请检查代理设置"
        if "timeout" in err.lower() or "timed out" in err.lower():
            return "请求超时，请重试"
        if "503" in err:
            return "苹果服务器繁忙，请稍后重试"
        if "541" in err:
            return "账号请求过于频繁，请稍后重试"
        if "401" in err:
            return "账号密码错误"
        if "404" in err:
            return "订单不存在"
        if "JSONDecodeError" in err or "json" in err.lower():
            return "页面解析失败，请重试"
        if "SSLError" in err:
            return "SSL连接错误，请重试"
        if "no _cs" in err or "no_cs" in err:
            return "无法获取订单会话，请重试"
        if "no stk" in err:
            return "登录后页面获取失败，请重试"
        if "no ssi" in err:
            return "登录跳转失败，请检查订单状态"
        if "no init" in err:
            return "页面解析失败，请重试"
        if "no call" in err:
            return "登录信息获取失败，请重试"
        if "authX" in err:
            return "登录授权失败，请重试"
        if "SRP" in err:
            return "苹果账号登录失败，请检查账号密码"
        if "xc failed" in err:
            return "订单链接无效或已过期"
        return err

    def process_one(self, row_idx, row_data):
        proxy = self.make_proxy()
        # 检测代理 IP
        proxy_ip = ""
        try:
            import requests
            r = requests.get("https://api.ipify.org", proxies={"http": proxy, "https": proxy} if proxy else None, timeout=8, verify=False)
            proxy_ip = r.text.strip()
            self.proxy_ip_signal.emit(row_idx, proxy_ip)
            self.log_signal.emit(f"[Row {row_idx+1}] 代理 IP: {proxy_ip}")
        except Exception as e:
            self.log_signal.emit(f"[Row {row_idx+1}] 代理检测失败，请检查代理设置")
        try:
            if self.task == "link":
                from apple_srp import link_order
                result = link_order(
                    order_number=row_data["order_number"],
                    order_email=row_data["order_email"],
                    apple_id=row_data["apple_id"],
                    apple_password=row_data["apple_pwd"],
                    phone=row_data["phone"],
                    proxy=proxy,
                )
                if result.get("success"):
                    return "绑定成功", ""
                return "失败", self._friendly_error(result.get("error", "未知错误"))

            elif self.task == "cancel":
                from cancel_order import cancel_order
                result = cancel_order(
                    order_number=row_data["order_number"],
                    order_email=row_data["order_email"],
                    apple_id=row_data["apple_id"],
                    apple_password=row_data["apple_pwd"],
                    phone=row_data["phone"],
                    proxy=proxy,
                )
                if result.get("success"):
                    return "取消成功", ""
                return "取消失败", self._friendly_error(result.get("error", "未知错误"))

            elif self.task == "address":
                from update_address import update_address
                result = update_address(
                    order_number=row_data["order_number"],
                    order_email=row_data["order_email"],
                    apple_id=row_data["apple_id"],
                    apple_password=row_data["apple_pwd"],
                    new_address=self.new_address,
                    proxy=proxy,
                )
                if result.get("success"):
                    return "地址修改成功", ""
                return "地址修改失败", self._friendly_error(result.get("error", "未知错误"))

        except Exception as e:
            return "失败", self._friendly_error(str(e))


# ── 地址输入对话框 ───────────────────────────────────────────────
class AddressDialog(QDialog):
    def __init__(self, parent=None, addr=None):
        super().__init__(parent)
        self.setWindowTitle("添加/编辑地址")
        self.setMinimumWidth(400)
        self._build_ui(addr or {})

    def _build_ui(self, addr):
        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.fields = {}
        fields_def = [
            ("name",       "地址名称",  "例：公司地址、家庭地址"),
            ("firstName",  "名 (First Name)", ""),
            ("lastName",   "姓 (Last Name)",  ""),
            ("street",     "街道地址",  ""),
            ("street2",    "套间/单元(可选)", ""),
            ("city",       "城市",      ""),
            ("state",      "州 (缩写)", "例：NY, CA, TX"),
            ("postalCode", "邮编",      ""),
        ]
        for key, label, placeholder in fields_def:
            edit = QLineEdit(addr.get(key, ""))
            if placeholder:
                edit.setPlaceholderText(placeholder)
            form.addRow(label + ":", edit)
            self.fields[key] = edit

        layout.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def get_address(self):
        return {k: v.text().strip() for k, v in self.fields.items()}


# ── 导入对话框 ────────────────────────────────────────────────────
class ImportDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("导入订单")
        self.setMinimumWidth(520)
        self.orders = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # 说明
        hint = QLabel("每行一条订单，用空格分隔：xc链接  苹果ID  苹果密码  手机号  [地址名称(可选)]")
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #666; font-size: 11px;")
        layout.addWidget(hint)

        # 输入框
        self.text_edit = QTextEdit()
        self.text_edit.setPlaceholderText(
            "https://www.apple.com/xc/us/vieworder/W0000000000/example@email.com  appleid@example.com  YourPassword  1234567890\n"
            "https://www.apple.com/xc/us/vieworder/W0000000001/example2@email.com  appleid2@example.com  YourPassword2  0987654321")
        self.text_edit.setMinimumHeight(200)
        layout.addWidget(self.text_edit)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_parse = QPushButton("解析预览")
        btn_parse.clicked.connect(self._parse_preview)
        btn_layout.addWidget(btn_parse)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # 预览表格
        self.preview = QTableWidget(0, 5)
        self.preview.setHorizontalHeaderLabels(["XC链接", "订单号", "订单邮箱", "苹果ID", "苹果密码", ])
        self.preview.horizontalHeader().setStretchLastSection(True)
        self.preview.setMaximumHeight(150)
        layout.addWidget(self.preview)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _parse_line(self, line):
        """解析一行：xc链接 苹果ID 苹果密码 手机号 [地址名称]（空格分隔）"""
        parts = line.split()
        xc = parts[0] if parts else ""
        apple_id    = parts[1] if len(parts) > 1 else ""
        apple_pwd   = parts[2] if len(parts) > 2 else ""
        phone       = parts[3] if len(parts) > 3 else ""
        new_address = " ".join(parts[4:]) if len(parts) > 4 else ""

        # 从 xc 链接解析订单号和邮箱
        import re
        m = re.search(r"/vieworder/([^/]+)/([^/\s?]+)", xc)
        order_number = m.group(1) if m else ""
        order_email  = m.group(2) if m else ""

        return {
            "xc_link":      xc,
            "order_number": order_number,
            "order_email":  order_email,
            "apple_id":     apple_id,
            "apple_pwd":    apple_pwd,
            "phone":        phone,
            "new_address":  new_address,
            "status":       "待处理",
            "note":         "",
        }

    def _parse_preview(self):
        lines = [l.strip() for l in self.text_edit.toPlainText().splitlines() if l.strip()]
        self.orders = [self._parse_line(l) for l in lines]
        self.preview.setRowCount(len(self.orders))
        for i, o in enumerate(self.orders):
            self.preview.setItem(i, 0, QTableWidgetItem(o["xc_link"][:40]))
            self.preview.setItem(i, 1, QTableWidgetItem(o["order_number"]))
            self.preview.setItem(i, 2, QTableWidgetItem(o["order_email"]))
            self.preview.setItem(i, 3, QTableWidgetItem(o["apple_id"]))
            self.preview.setItem(i, 4, QTableWidgetItem(o["apple_pwd"]))

    def get_orders(self):
        if not self.orders:
            self._parse_preview()
        return self.orders

# ── 代理设置对话框 ────────────────────────────────────────────────
class ProxyDialog(QDialog):
    def __init__(self, cfg, parent=None):
        super().__init__(parent)
        self.setWindowTitle("代理设置")
        self.setMinimumWidth(400)
        self.cfg = cfg.copy()

        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.enabled_cb = QCheckBox("启用代理")
        self.enabled_cb.setChecked(cfg.get("enabled", False))
        form.addRow("", self.enabled_cb)

        self.type_combo = QComboBox()
        self.type_combo.addItems(["Smartproxy", "Bright Data"])
        self.type_combo.setCurrentText(cfg.get("type", "Smartproxy"))
        form.addRow("代理类型:", self.type_combo)

        self.host_edit = QLineEdit(cfg.get("host", "proxy.smartproxy.net"))
        form.addRow("主机:", self.host_edit)

        self.port_edit = QLineEdit(cfg.get("port", "3120"))
        form.addRow("端口:", self.port_edit)

        self.user_edit = QLineEdit(cfg.get("user", ""))
        form.addRow("用户名:", self.user_edit)

        self.pwd_edit = QLineEdit(cfg.get("pwd", ""))
        self.pwd_edit.setEchoMode(QLineEdit.Password)
        form.addRow("密码:", self.pwd_edit)

        layout.addLayout(form)

        # 测试代理按钮
        test_layout = QHBoxLayout()
        self.btn_test = QPushButton("测试代理")
        self.btn_test.clicked.connect(self._test_proxy)
        self.test_result = QLabel("未测试")
        self.test_result.setStyleSheet("color: #666;")
        test_layout.addWidget(self.btn_test)
        test_layout.addWidget(self.test_result)
        test_layout.addStretch()
        layout.addLayout(test_layout)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _test_proxy(self):
        cfg = self.get_config()
        if not cfg.get("enabled") or not cfg.get("host"):
            self.test_result.setText("❌ 请先填写代理信息")
            self.test_result.setStyleSheet("color: red;")
            return
        self.btn_test.setEnabled(False)
        self.test_result.setText("测试中...")
        self.test_result.setStyleSheet("color: #666;")

        import threading, time, requests
        def do_test():
            try:
                import random, string
                sid = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
                user = cfg.get("user", "")
                pwd = cfg.get("pwd", "")
                host = cfg.get("host", "")
                port = cfg.get("port", "3120")
                proxy_type = cfg.get("type", "Smartproxy")
                if proxy_type == "Bright Data":
                    proxy_url = f"http://{user}-session-{sid}:{pwd}@{host}:{port}"
                else:
                    proxy_url = f"http://{user}_session-{sid}:{pwd}@{host}:{port}"
                proxies = {"http": proxy_url, "https": proxy_url}
                start = time.time()
                r = requests.get("https://api.ipify.org", proxies=proxies,
                                timeout=10, verify=False)
                elapsed = int((time.time() - start) * 1000)
                ip = r.text.strip()
                self.test_result.setText(f"✅ {ip} | {elapsed}ms")
                self.test_result.setStyleSheet("color: green;")
            except Exception as e:
                self.test_result.setText(f"❌ 连接失败")
                self.test_result.setStyleSheet("color: red;")
            finally:
                self.btn_test.setEnabled(True)

        threading.Thread(target=do_test, daemon=True).start()

    def get_config(self):
        return {
            "enabled": self.enabled_cb.isChecked(),
            "type":    self.type_combo.currentText(),
            "host":    self.host_edit.text().strip(),
            "port":    self.port_edit.text().strip(),
            "user":    self.user_edit.text().strip(),
            "pwd":     self.pwd_edit.text().strip(),
        }

# ── 主窗口 ────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"极限苹果订单修改管理系统 v{APP_VERSION}")
        self.setMinimumSize(1280, 720)
        self.proxy_cfg = {"enabled": False, "host": "", "port": "3120", "user": "", "pwd": ""}
        self.worker = None
        self._load_proxy_cfg()
        self._build_ui()
        # 启动后检查更新
        QTimer.singleShot(2000, self._check_update)

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)

        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)
        main_layout.addWidget(self.tabs)

        # Tab 1: 订单管理
        self.order_tab = QWidget()
        self.tabs.addTab(self.order_tab, "📋 订单管理")
        order_layout = QVBoxLayout(self.order_tab)
        order_layout.setSpacing(8)
        order_layout.setContentsMargins(10, 10, 10, 10)

        # Tab 2: 订单查询
        self.query_tab = QWidget()
        self.tabs.addTab(self.query_tab, "🔍 订单查询")
        self._build_query_tab()

        # Tab 3: 地址管理
        self.address_tab = QWidget()
        self.tabs.addTab(self.address_tab, "📮 地址管理")
        self._build_address_tab()

        # 后面的代码都放进 order_layout
        central2 = self.order_tab
        main_layout2 = order_layout

        # ── 顶部工具栏 ─────────────────────────────────────────────
        toolbar = QHBoxLayout()

        self.btn_import = QPushButton("📂 导入")
        self.btn_import.setToolTip("从 CSV/Excel 导入订单")
        self.btn_import.clicked.connect(self.import_orders)

        self.btn_add_row = QPushButton("➕ 添加行")
        self.btn_add_row.clicked.connect(self.add_row)

        self.btn_del_row = QPushButton("🗑 删除行")
        self.btn_del_row.clicked.connect(self.del_row)

        self.btn_export = QPushButton("💾 导出结果")
        self.btn_export.clicked.connect(self.export_results)

        self.btn_proxy = QPushButton("🌐 代理设置")
        self.btn_proxy.clicked.connect(self.open_proxy)

        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)

        sep2 = QFrame()
        sep2.setFrameShape(QFrame.VLine)

        self.btn_link   = QPushButton("🔗 绑定订单")
        self.btn_cancel = QPushButton("❌ 取消订单")
        self.btn_addr   = QPushButton("📦 修改地址")
        self.btn_stop   = QPushButton("⏹ 停止")
        self.btn_stop.setEnabled(False)

        self.btn_link.clicked.connect(lambda: self.run_task("link"))
        self.btn_cancel.clicked.connect(lambda: self.run_task("cancel"))
        self.btn_addr.clicked.connect(lambda: self.run_task("address"))
        self.btn_stop.clicked.connect(self.stop_task)

        for btn in [self.btn_link, self.btn_cancel, self.btn_addr]:
            btn.setStyleSheet("QPushButton { background: #007AFF; color: white; border-radius: 4px; padding: 5px 12px; }"
                              "QPushButton:hover { background: #0051D5; }")

        self.btn_stop.setStyleSheet("QPushButton { background: #FF3B30; color: white; border-radius: 4px; padding: 5px 12px; }"
                                    "QPushButton:disabled { background: #ccc; }")

        for w in [self.btn_import, self.btn_add_row, self.btn_del_row,
                  self.btn_export, self.btn_proxy]:
            toolbar.addWidget(w)
        toolbar.addWidget(sep)
        toolbar.addWidget(sep2)
        toolbar.addWidget(self.btn_link)
        toolbar.addWidget(self.btn_cancel)
        toolbar.addWidget(self.btn_addr)
        toolbar.addWidget(self.btn_stop)
        toolbar.addStretch()

        order_layout.addLayout(toolbar)

        # ── 进度条 ─────────────────────────────────────────────────
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setTextVisible(True)
        order_layout.addWidget(self.progress)

        # ── Splitter: 表格 + 日志 ──────────────────────────────────
        splitter = QSplitter(Qt.Vertical)

        # 表格
        self.table = QTableWidget()
        self.table.setColumnCount(len(COLUMNS))
        self.table.setHorizontalHeaderLabels([label for _, label in COLUMNS])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setDefaultSectionSize(28)
        self.table.setColumnWidth(COL_IDX["xc_link"],     200)
        self.table.setColumnWidth(COL_IDX["order_number"],120)
        self.table.setColumnWidth(COL_IDX["order_email"], 180)
        self.table.setColumnWidth(COL_IDX["apple_id"],    180)
        self.table.setColumnWidth(COL_IDX["apple_pwd"],   120)
        self.table.setColumnWidth(COL_IDX["phone"],       110)
        self.table.setColumnWidth(COL_IDX["new_address"], 180)
        self.table.setColumnWidth(COL_IDX["proxy_ip"],    120)
        self.table.setColumnWidth(COL_IDX["status"],       90)
        splitter.addWidget(self.table)

        # 日志
        log_widget = QWidget()
        log_layout = QVBoxLayout(log_widget)
        log_layout.setContentsMargins(0, 0, 0, 0)
        log_label = QLabel("日志")
        log_label.setFont(QFont("Arial", 9))
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setMaximumHeight(160)
        self.log_box.setFont(QFont("Courier", 9))
        self.log_box.setStyleSheet("background: #1E1E1E; color: #D4D4D4;")
        log_layout.addWidget(log_label)
        log_layout.addWidget(self.log_box)
        splitter.addWidget(log_widget)
        splitter.setSizes([520, 160])

        order_layout.addWidget(splitter)

        # ── 状态栏 ─────────────────────────────────────────────────
        self.statusBar().showMessage("就绪")

        # 不添加初始空行，等用户导入

    # ── 表格操作 ──────────────────────────────────────────────────
    def add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        for col in range(len(COLUMNS)):
            item = QTableWidgetItem("")
            self.table.setItem(row, col, item)
        self._set_status(row, "待处理", "")

    def del_row(self):
        rows = sorted(set(i.row() for i in self.table.selectedItems()), reverse=True)
        for row in rows:
            self.table.removeRow(row)

    def _set_status(self, row, status, note=""):
        status_item = QTableWidgetItem(status)
        color = STATUS_COLORS.get(status, "#FFFFFF")
        status_item.setBackground(QColor(color))
        status_item.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(row, COL_IDX["status"], status_item)
        note_item = QTableWidgetItem(note)
        self.table.setItem(row, COL_IDX["note"], note_item)

    def _get_row_data(self, row):
        def cell(col_key):
            item = self.table.item(row, COL_IDX[col_key])
            return item.text().strip() if item else ""
        return {k: cell(k) for k, _ in COLUMNS}

    # ── 导入 ──────────────────────────────────────────────────────
    def _build_address_tab(self):
        layout = QVBoxLayout(self.address_tab)
        layout.setSpacing(8)
        layout.setContentsMargins(10, 10, 10, 10)

        # 工具栏
        toolbar = QHBoxLayout()
        btn_style = "QPushButton { border-radius: 4px; padding: 5px 12px; }"
        btn_add_addr = QPushButton("添加地址")
        btn_add_addr.setStyleSheet("QPushButton { background: #007AFF; color: white; border-radius: 4px; padding: 5px 12px; }"
                                   "QPushButton:hover { background: #0051D5; }")
        btn_add_addr.clicked.connect(self._add_address)
        btn_edit_addr = QPushButton("编辑地址")
        btn_edit_addr.setStyleSheet(btn_style)
        btn_edit_addr.clicked.connect(self._edit_address)
        btn_del_addr = QPushButton("删除地址")
        btn_del_addr.setStyleSheet(btn_style)
        btn_del_addr.clicked.connect(self._del_address)
        toolbar.addWidget(btn_add_addr)
        toolbar.addWidget(btn_edit_addr)
        toolbar.addWidget(btn_del_addr)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        # 地址列表
        ADDR_COLS = ["name", "firstName", "lastName", "street", "street2", "city", "state", "postalCode"]
        ADDR_LABELS = ["地址名称", "名", "姓", "街道", "套间/单元", "城市", "州", "邮编"]
        self.addr_col_idx = {k: i for i, k in enumerate(ADDR_COLS)}

        self.addr_table = QTableWidget()
        self.addr_table.setColumnCount(len(ADDR_COLS))
        self.addr_table.setHorizontalHeaderLabels(ADDR_LABELS)
        self.addr_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.addr_table.horizontalHeader().setStretchLastSection(True)
        self.addr_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.addr_table.setAlternatingRowColors(True)
        self.addr_table.verticalHeader().setDefaultSectionSize(28)
        self.addr_table.setColumnWidth(0, 120)
        self.addr_table.setColumnWidth(1, 80)
        self.addr_table.setColumnWidth(2, 80)
        self.addr_table.setColumnWidth(3, 180)
        self.addr_table.setColumnWidth(4, 100)
        self.addr_table.setColumnWidth(5, 100)
        self.addr_table.setColumnWidth(6, 50)
        self.addr_table.setColumnWidth(7, 80)
        layout.addWidget(self.addr_table)

        # 加载保存的地址
        self._load_addresses()

    def _get_address_names(self):
        names = []
        for row in range(self.addr_table.rowCount()):
            item = self.addr_table.item(row, self.addr_col_idx["name"])
            names.append(item.text() if item else f"地址{row+1}")
        return names

    def _get_address_by_name(self, name):
        for row in range(self.addr_table.rowCount()):
            item = self.addr_table.item(row, self.addr_col_idx["name"])
            if item and item.text() == name:
                return {k: (self.addr_table.item(row, i).text()
                           if self.addr_table.item(row, i) else "")
                       for k, i in self.addr_col_idx.items()}
        return None

    def _load_addresses(self):
        import os, json
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "addresses.json")
        if not os.path.exists(path):
            return
        try:
            with open(path, 'r') as f:
                addresses = json.load(f)
            for addr in addresses:
                self._insert_address_row(addr)
        except Exception:
            pass

    def _save_addresses(self):
        import os, json
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "addresses.json")
        addresses = []
        for row in range(self.addr_table.rowCount()):
            addr = {k: (self.addr_table.item(row, i).text()
                       if self.addr_table.item(row, i) else "")
                   for k, i in self.addr_col_idx.items()}
            addresses.append(addr)
        with open(path, 'w') as f:
            json.dump(addresses, f, ensure_ascii=False, indent=2)

    def _insert_address_row(self, addr=None):
        row = self.addr_table.rowCount()
        self.addr_table.insertRow(row)
        for key, idx in self.addr_col_idx.items():
            val = addr.get(key, "") if addr else ""
            self.addr_table.setItem(row, idx, QTableWidgetItem(val))

    def _add_address(self):
        dlg = AddressDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            self._insert_address_row(dlg.get_address())
            self._save_addresses()

    def _edit_address(self):
        rows = list(set(i.row() for i in self.addr_table.selectedItems()))
        if not rows:
            QMessageBox.warning(self, "提示", "请先选择一个地址")
            return
        row = rows[0]
        addr = {k: (self.addr_table.item(row, i).text()
                   if self.addr_table.item(row, i) else "")
               for k, i in self.addr_col_idx.items()}
        dlg = AddressDialog(self, addr)
        if dlg.exec_() == QDialog.Accepted:
            new_addr = dlg.get_address()
            for key, idx in self.addr_col_idx.items():
                self.addr_table.setItem(row, idx, QTableWidgetItem(new_addr.get(key, "")))
            self._save_addresses()

    def _del_address(self):
        rows = sorted(set(i.row() for i in self.addr_table.selectedItems()), reverse=True)
        if not rows:
            QMessageBox.warning(self, "提示", "请先选择地址")
            return
        for row in rows:
            self.addr_table.removeRow(row)
        self._save_addresses()

    def _build_query_tab(self):
        layout = QVBoxLayout(self.query_tab)
        layout.setSpacing(8)
        layout.setContentsMargins(10, 10, 10, 10)

        # 工具栏
        query_toolbar = QHBoxLayout()

        btn_import_query = QPushButton("📂 导入链接")
        btn_import_query.clicked.connect(self._import_query_links)

        btn_query = QPushButton("🔍 查询")
        btn_query.setStyleSheet("QPushButton { background: #007AFF; color: white; border-radius: 4px; padding: 5px 12px; }"
                                "QPushButton:hover { background: #0051D5; }")
        btn_query.clicked.connect(self._run_query)

        btn_clear = QPushButton("🗑 清空")
        btn_clear.clicked.connect(self._clear_query)

        query_toolbar.addWidget(btn_import_query)
        query_toolbar.addWidget(btn_query)
        query_toolbar.addWidget(btn_clear)
        query_toolbar.addStretch()
        layout.addLayout(query_toolbar)

        # 进度条
        self.query_progress = QProgressBar()
        self.query_progress.setVisible(False)
        layout.addWidget(self.query_progress)

        # 结果表格
        QUERY_COLS = ["xc_link", "order_number", "order_email", "order_status", "order_date", "device", "tracking", "carrier", "eta", "address"]
        QUERY_LABELS = ["XC 链接", "订单号", "订单邮箱", "订单状态", "下单日期", "购买设备", "物流号", "物流公司", "预计到达", "收货地址"]
        self.query_col_idx = {k: i for i, k in enumerate(QUERY_COLS)}

        self.query_table = QTableWidget()
        self.query_table.setColumnCount(len(QUERY_COLS))
        self.query_table.setHorizontalHeaderLabels(QUERY_LABELS)
        self.query_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.query_table.horizontalHeader().setStretchLastSection(True)
        self.query_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.query_table.setAlternatingRowColors(True)
        self.query_table.verticalHeader().setDefaultSectionSize(28)
        self.query_table.setColumnWidth(0, 210)
        self.query_table.setColumnWidth(1, 120)
        self.query_table.setColumnWidth(2, 160)
        self.query_table.setColumnWidth(3, 90)
        self.query_table.setColumnWidth(4, 100)
        self.query_table.setColumnWidth(5, 160)
        self.query_table.setColumnWidth(6, 130)
        self.query_table.setColumnWidth(7, 70)
        self.query_table.setColumnWidth(8, 110)
        self.query_table.setColumnWidth(9, 180)
        layout.addWidget(self.query_table)

        # 导出按钮
        btn_export_query = QPushButton("💾 导出查询结果")
        btn_export_query.clicked.connect(self._export_query)
        layout.addWidget(btn_export_query)

    def _import_query_links(self):
        """弹出输入框，粘贴 xc 链接导入"""
        dlg = QDialog(self)
        dlg.setWindowTitle("导入 XC 链接")
        dlg.setMinimumWidth(500)
        layout = QVBoxLayout(dlg)

        hint = QLabel("每行一个 xc 链接：")
        hint.setStyleSheet("color: #666; font-size: 11px;")
        layout.addWidget(hint)

        text_edit = QTextEdit()
        text_edit.setPlaceholderText(
            "https://www.apple.com/xc/us/vieworder/W0000000000/example@email.com\n"
            "https://www.apple.com/xc/us/vieworder/W0000000001/example2@email.com")
        text_edit.setMinimumHeight(200)
        layout.addWidget(text_edit)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        layout.addWidget(btns)

        if dlg.exec_() == QDialog.Accepted:
            lines = [l.strip() for l in text_edit.toPlainText().splitlines()
                     if l.strip() and "vieworder" in l]
            if lines:
                # 加到查询表格
                for link in lines:
                    row = self.query_table.rowCount()
                    self.query_table.insertRow(row)
                    item = QTableWidgetItem(link)
                    self.query_table.setItem(row, self.query_col_idx["xc_link"], item)
                    self.query_table.setItem(row, self.query_col_idx["order_status"],
                                             QTableWidgetItem("待查询"))
                self._log(f"导入 {len(lines)} 条查询链接")

    def _run_query(self):
        # 查所有行
        lines = []
        for row in range(self.query_table.rowCount()):
            xc_item = self.query_table.item(row, self.query_col_idx["xc_link"])
            if xc_item and xc_item.text().strip():
                lines.append(xc_item.text().strip())
        if not lines:
            QMessageBox.warning(self, "提示", "请先导入 xc 链接")
            return
        self.query_progress.setVisible(True)
        self.query_progress.setMaximum(len(lines))
        self.query_progress.setValue(0)
        self._query_worker = QueryWorker(lines, self.proxy_cfg)
        self._query_worker.row_ready.connect(self._on_query_row)
        self._query_worker.finished.connect(self._on_query_finished)
        self._query_worker.start()

    def _on_query_row(self, data):
        # 找到对应的行更新
        xc = data.get("xc_link", "")
        for row in range(self.query_table.rowCount()):
            item = self.query_table.item(row, self.query_col_idx["xc_link"])
            if item and item.text().strip() == xc:
                for key, val in data.items():
                    if key in self.query_col_idx:
                        cell = QTableWidgetItem(val)
                        cell.setTextAlignment(Qt.AlignCenter)
                        self.query_table.setItem(row, self.query_col_idx[key], cell)
                break
        self.query_progress.setValue(self.query_progress.value() + 1)

    def _on_query_finished(self):
        self.query_progress.setVisible(False)
        self._auto_export_query()

    def _auto_export_query(self):
        if not HAS_OPENPYXL:
            return
        try:
            import os
            ts = datetime.now().strftime("%H%M%S")
            date_str = datetime.now().strftime("%Y%m%d")
            base_dir = os.path.dirname(os.path.abspath(__file__))
            day_dir = os.path.join(base_dir, "results", date_str)
            os.makedirs(day_dir, exist_ok=True)
            path = os.path.join(day_dir, f"查询结果_{ts}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "查询结果"
            labels = ["XC 链接", "订单号", "订单邮箱", "订单状态", "下单日期", "购买设备", "物流号", "物流公司", "预计到达", "收货地址"]
            ws.append(labels)
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row in range(self.query_table.rowCount()):
                data = []
                for col in range(self.query_table.columnCount()):
                    item = self.query_table.item(row, col)
                    data.append(item.text() if item else "")
                ws.append(data)
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
            wb.save(path)
            self._log(f"📊 查询结果已保存: {path}")
        except Exception as e:
            self._log(f"查询结果保存失败: {e}")

    def _clear_query(self):
        self.query_table.setRowCount(0)

    def _export_query(self):
        if not HAS_OPENPYXL:
            QMessageBox.critical(self, "错误", "请安装 openpyxl")
            return
        import os
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        date_str = datetime.now().strftime("%Y%m%d")
        base_dir = os.path.dirname(os.path.abspath(__file__))
        day_dir = os.path.join(base_dir, "results", date_str)
        os.makedirs(day_dir, exist_ok=True)
        path = os.path.join(day_dir, f"查询结果_{ts}.xlsx")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            labels = ["XC 链接", "订单号", "订单状态", "下单日期", "购买设备", "金额"]
            ws.append(labels)
            for row in range(self.query_table.rowCount()):
                data = []
                for col in range(self.query_table.columnCount()):
                    item = self.query_table.item(row, col)
                    data.append(item.text() if item else "")
                ws.append(data)
            wb.save(path)
            self._log(f"查询结果已导出: {path}")
            QMessageBox.information(self, "导出成功", f"已保存到:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def import_orders(self):
        dlg = ImportDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return
        orders = dlg.get_orders()
        if not orders:
            return
        for o in orders:
            row = self.table.rowCount()
            self.table.insertRow(row)
            for col_key, _ in COLUMNS:
                item = QTableWidgetItem(o.get(col_key, ""))
                self.table.setItem(row, COL_IDX[col_key], item)
            self._set_status(row, "待处理", "")
        self._log(f"导入 {len(orders)} 条订单")

    def _read_file(self, path):
        if path.endswith('.csv'):
            import csv
            with open(path, encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                return list(reader)
        else:
            if not HAS_OPENPYXL:
                raise Exception("请安装 openpyxl: pip install openpyxl")
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = {headers[i]: (str(v).strip() if v else "") for i, v in enumerate(row)}
                rows.append(d)
            return rows

    # ── 导出 ──────────────────────────────────────────────────────
    def export_results(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "导出结果", f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel (*.xlsx)")
        if not path:
            return
        if not HAS_OPENPYXL:
            QMessageBox.critical(self, "错误", "请安装 openpyxl")
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "结果"
            headers = [label for _, label in COLUMNS]
            ws.append(headers)
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row in range(self.table.rowCount()):
                data = []
                for col in range(len(COLUMNS)):
                    item = self.table.item(row, col)
                    data.append(item.text() if item else "")
                ws.append(data)
                status = data[COL_IDX["status"]]
                if "成功" in status:
                    fill = PatternFill("solid", fgColor="C6EFCE")
                elif "失败" in status:
                    fill = PatternFill("solid", fgColor="FFC7CE")
                else:
                    fill = None
                if fill:
                    for cell in ws[row + 2]:
                        cell.fill = fill
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            wb.save(path)
            self._log(f"导出成功: {path}")
            QMessageBox.information(self, "导出成功", f"已保存到:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    # ── 代理 ──────────────────────────────────────────────────────
    def open_proxy(self):
        dlg = ProxyDialog(self.proxy_cfg, self)
        if dlg.exec_() == QDialog.Accepted:
            self.proxy_cfg = dlg.get_config()
            self._save_proxy_cfg()
            status = "已启用" if self.proxy_cfg["enabled"] else "未启用"
            self._log(f"代理设置已更新并保存（{status}）")

    # ── 任务执行 ──────────────────────────────────────────────────
    def run_task(self, task):
        # 收集需要处理的行
        selected = [i.row() for i in self.table.selectedItems()]
        selected = sorted(set(selected))
        if not selected:
            selected = list(range(self.table.rowCount()))
        if not selected:
            QMessageBox.warning(self, "提示", "没有可处理的订单")
            return

        rows = []
        for row in selected:
            data = self._get_row_data(row)
            if not data["order_number"]:
                continue
            rows.append((row, data))

        if not rows:
            QMessageBox.warning(self, "提示", "没有有效的订单行")
            return

        task_name = {"link": "绑定订单", "cancel": "取消订单", "address": "修改地址"}[task]

        # 修改地址 - 弹出地址选择框
        selected_address = None
        if task == "address":
            addr_names = [n for n in self._get_address_names() if n.strip()]
            if not addr_names:
                QMessageBox.warning(self, "提示", "请先在【地址管理】页面添加地址")
                return
            dlg = QDialog(self)
            dlg.setWindowTitle("选择新地址")
            dlg.setMinimumWidth(320)
            dlg_layout = QVBoxLayout(dlg)
            dlg_layout.addWidget(QLabel("请选择要修改的目标地址："))
            combo = QComboBox()
            combo.addItems(addr_names)
            dlg_layout.addWidget(combo)
            btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            btns.accepted.connect(dlg.accept)
            btns.rejected.connect(dlg.reject)
            dlg_layout.addWidget(btns)
            if dlg.exec_() != QDialog.Accepted:
                return
            selected_addr_name = combo.currentText()
            selected_address = self._get_address_by_name(selected_addr_name)
            # 把选中的地址名称显示在表格的修改地址列
            for row_idx, _ in rows:
                self.table.setItem(row_idx, COL_IDX["new_address"],
                                   QTableWidgetItem(selected_addr_name))

        reply = QMessageBox.question(
            self, "确认", f"将对 {len(rows)} 个订单执行【{task_name}】操作，确认？",
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.progress.setVisible(True)
        self.progress.setMaximum(len(rows))
        self.progress.setValue(0)
        self._completed = 0

        self.btn_link.setEnabled(False)
        self.btn_cancel.setEnabled(False)
        self.btn_addr.setEnabled(False)
        self.btn_stop.setEnabled(True)

        self.worker = WorkerThread(task, rows, self.proxy_cfg,
                                   len(rows), selected_address)  # 自动并发，导入几个跑几个
        self.worker.row_update.connect(self._on_row_update)
        self.worker.proxy_ip_signal.connect(self._on_proxy_ip)
        self.worker.log_signal.connect(self._log)
        self.worker.finished.connect(self._on_finished)
        self.worker.start()
        self._log(f"开始执行【{task_name}】，共 {len(rows)} 个订单")

    def stop_task(self):
        if self.worker:
            self.worker.stop()
            self._log("正在停止...")
            self.btn_stop.setEnabled(False)

    def _on_proxy_ip(self, row, ip):
        item = QTableWidgetItem(ip)
        item.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(row, COL_IDX["proxy_ip"], item)

    def _on_row_update(self, row, status, note):
        self._set_status(row, status, note)
        if status != "处理中":
            self._completed += 1
            self.progress.setValue(self._completed)
            self._log(f"[{self.table.item(row, COL_IDX['order_number']).text()}] {status} {note}")

    def _on_finished(self):
        self.btn_link.setEnabled(True)
        self.btn_cancel.setEnabled(True)
        self.btn_addr.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.progress.setVisible(False)
        success = sum(1 for r in range(self.table.rowCount())
                      if "成功" in (self.table.item(r, COL_IDX["status"]) or QTableWidgetItem("")).text())
        self._log(f"✅ 任务完成，成功 {success} 个")
        self.statusBar().showMessage(f"完成，成功 {success} 个")
        # 自动保存 Excel
        self._auto_export()

    def _auto_export(self):
        if not HAS_OPENPYXL:
            return
        try:
            import os
            ts = datetime.now().strftime("%H%M%S")
            date_str = datetime.now().strftime("%Y%m%d")
            # 项目目录下按日期建文件夹
            base_dir = os.path.dirname(os.path.abspath(__file__))
            day_dir = os.path.join(base_dir, "results", date_str)
            os.makedirs(day_dir, exist_ok=True)
            path = os.path.join(day_dir, f"订单结果_{ts}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "结果"
            headers = [label for _, label in COLUMNS]
            ws.append(headers)
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row in range(self.table.rowCount()):
                data = []
                for col in range(len(COLUMNS)):
                    item = self.table.item(row, col)
                    data.append(item.text() if item else "")
                ws.append(data)
                status = data[COL_IDX["status"]]
                if "成功" in status:
                    fill = PatternFill("solid", fgColor="C6EFCE")
                elif "失败" in status:
                    fill = PatternFill("solid", fgColor="FFC7CE")
                else:
                    fill = None
                if fill:
                    for cell in ws[row + 2]:
                        cell.fill = fill
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            wb.save(path)
            self._log(f"📊 结果已自动保存: {path}")
        except Exception as e:
            self._log(f"自动保存失败: {e}")

    def _check_update(self):
        import threading
        def do_check():
            try:
                import urllib.request
                latest = urllib.request.urlopen(VERSION_URL, timeout=5).read().decode().strip()
                if latest != APP_VERSION:
                    self._latest_version = latest
                    QTimer.singleShot(0, self._show_update_dialog)
            except Exception:
                pass
        threading.Thread(target=do_check, daemon=True).start()

    def _show_update_dialog(self):
        latest_version = getattr(self, "_latest_version", "")
        msg = f"当前版本：{APP_VERSION}\n最新版本：{latest_version}\n\n是否下载更新？"
        reply = QMessageBox.question(
            self, "发现新版本", msg,
            QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self._download_update()

    def _download_update(self):
        import webbrowser
        webbrowser.open(DOWNLOAD_URL)
        self._log(f"正在打开下载页面，下载完成后替换程序即可")

    def _load_proxy_cfg(self):
        import os, json
        cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "proxy_cfg.json")
        try:
            if os.path.exists(cfg_path):
                with open(cfg_path, 'r') as f:
                    self.proxy_cfg = json.load(f)
        except Exception:
            pass

    def _save_proxy_cfg(self):
        import os, json
        cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "proxy_cfg.json")
        try:
            with open(cfg_path, 'w') as f:
                json.dump(self.proxy_cfg, f, indent=2)
        except Exception:
            pass

    def _log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.append(f"[{ts}] {msg}")
        self.log_box.verticalScrollBar().setValue(
            self.log_box.verticalScrollBar().maximum())


# ── 入口 ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())